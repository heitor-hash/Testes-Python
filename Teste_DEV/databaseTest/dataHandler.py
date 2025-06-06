# from openpyxl.utils import column_index_from_string
from openpyxl.utils import get_column_letter
import openpyxl as xl
import easygui as egui
import os


def open_workbook(path:str) -> xl.Workbook:
    try:
        wb = xl.load_workbook(filename=path)
        print("Workbook loaded")
    except:
        wb = xl.Workbook()
    return wb
    
def open_sheet(wb:xl.Workbook, sheetname: str):
    try:
        sheet = wb[sheetname]
    except:
        sheet = wb.create_sheet(sheetname)
    return sheet

def excel_save_tuple_at_index(sheet: str, index:int, item_tuple):
    a = get_column_letter(index)
    for cell in sheet[f'{a}']:
        cell = None
    i = 1
    for item in item_tuple:
        sheet[f'{a}{i}'] = item
        i += 1
    return
    

def get_excel_path() -> str:
    path = egui.fileopenbox(default='*.xlsx',
    title='Abrir arquivo Excel',
    filetypes=['*.xlsx']
    )
    r, ext = os.path.splitext(path)
    if ext != '.xlsx':
        return None
    else:
        return path
    

def save_excel_multiple_tuples(sheetname:str, *args):
    path = egui.filesavebox(default='*.xlsx',
    title='Salvar Como')
    
    r, ext = os.path.splitext(path)
    path = r + '.xlsx'
    
    wb = open_workbook(path=path)
    sheet = open_sheet(wb=wb, sheetname=sheetname)
    i = 1
    for arg in args:
        excel_save_tuple_at_index(sheet, i, arg)
        i += 1
    wb.save()
    wb.close()
    
def excel_data_to_matrix(sheet):
    matrix = []
    i = 0
    for col in sheet['1']:
        print(sheet['1'])
        
        column_letter = get_column_letter(i+1)
        internal_matrix = []
        for cell in sheet[f'{column_letter}']:
            internal_matrix.append(cell.value)
        matrix.append(internal_matrix)
        i += 1
    print(f"MATRIX {matrix}")
    return matrix
        