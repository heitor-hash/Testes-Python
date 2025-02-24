import easygui
import os
import openpyxl as xl
from openpyxl.utils import column_index_from_string
from openpyxl.utils import get_column_letter




# try to open a workbook
def open_workbook(filename:str) -> xl.Workbook:
    try:
        wb = xl.load_workbook(filename=filename)
        print("WORK")
        return wb
    except:
        raise FileNotFoundError
    
def new_workbook(filename:str) -> xl.Workbook:
    try:
        wb = xl.load_workbook(filename=filename)
        print("WORK")
        return wb
    except:
        wb = xl.Workbook()
        return wb
    

def worksheet_to_dict(wb: xl.Workbook, sheet: str) -> dict | None:
    ws = wb[sheet]
    
    keys = []
    for cell in ws['1']:
        keys.append(cell.value)
    values = []
    for cell in ws['2']:
        values.append(cell.value)
        
    dicio = dict(zip(keys, values))
    return dicio

def dict_to_worksheet(wb: xl.Workbook, sheet: str, dictionary: dict, filename: str) -> None:
    try:
        ws = wb[sheet]
    except:
        ws = wb.create_sheet(title=sheet)
    keys = dictionary.keys()
    values = dictionary.values()
    
    for cells in ws[1:2]:
        for cell in cells:
            cell.value = None
    
    # For key in keys add value to 'KeyIndex 1'
    index = 1
    for key in keys:
        a = get_column_letter(index)
        ws[f'{a}1'] = key
        index += 1
    index = 1
    for value in values:
        a = get_column_letter(index)
        ws[f'{a}2'] = value
        index += 1
    
    wb.save(filename=filename)

def easy_xl_file_selector() -> str:
    item = easygui.fileopenbox(
        default='*.xlsx',
        title='Selecione arquivo Excel para dados',
        filetypes=['*.xlsx']
        )
    print(f"O arquivo é: {item}")
    return item
    
def easy_xl_save_file() -> str:
    name:str = easygui.filesavebox(default='*.xlsx',
                        title='Salvar como',# quero que sempre salve como arquivo xlsx
                        filetypes=['*.xlsx']
                        )
    if name:
        n, ext = os.path.splitext(name)
        name = n + '.xlsx'
    return name